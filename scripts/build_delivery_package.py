#!/usr/bin/env python3
"""Build a user-friendly Chinese delivery folder from a live-monitor run."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


SPEECH_TYPE_CN = {
    "product_intro": "商品介绍",
    "price_coupon": "价格/券/福利",
    "urgency": "紧迫感",
    "trust": "信任背书",
    "social_proof": "销量/人气证明",
    "objection_handling": "异议处理",
    "cta": "行动引导",
    "interaction": "互动",
    "scene_emotion": "场景情绪",
    "brand_story": "品牌故事",
    "operation": "直播操作",
}

EVENT_KIND_CN = {
    "product_switch": "商品切换",
    "price_change": "价格变化",
    "coupon_drop": "发券",
    "stock_claim": "库存话术",
    "limited_time": "限时",
    "bundle": "组合",
    "shipping": "发货",
    "gift": "赠品",
    "sold_count_change": "销量变化",
}

COMMENT_TYPE_CN = {
    "buyer_question": "购买问题",
    "price_question": "价格问题",
    "quality_question": "品质问题",
    "shipping_question": "发货问题",
    "coupon_question": "优惠问题",
    "social_proof": "官方/口碑回复",
    "complaint": "投诉",
    "greeting": "问候",
    "other": "其他",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-dir", required=True, help="Raw live-monitor run directory")
    parser.add_argument("--out-dir", help="Delivery directory. Default: sibling Chinese folder")
    parser.add_argument("--account-name", help="Override account name")
    parser.add_argument("--date", help="Override session date")
    parser.add_argument("--duration-label", help="Override duration label, e.g. 10分钟")
    return parser.parse_args()


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    records = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            records.append(json.loads(line))
    return records


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def safe_filename(text: str, max_len: int = 70) -> str:
    text = re.sub(r"[\\/:*?\"<>|]+", " ", text or "")
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace("\n", " ")
    return text[:max_len].strip(" ._") or "未命名"


def iso_to_clock(value: str | None) -> str:
    if not value:
        return ""
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S.%f%z"):
        try:
            dt = datetime.strptime(value.replace("+08:00", "+0800"), fmt)
            return dt.strftime("%H:%M:%S")
        except ValueError:
            pass
    return value


def timecode_to_cn(value: str | None) -> str:
    if not value:
        return "未知时间"
    parts = value.split(":")
    try:
        if len(parts) == 3:
            h, m, s = [int(float(p)) for p in parts]
        elif len(parts) == 2:
            h, m, s = 0, int(float(parts[0])), int(float(parts[1]))
        else:
            total = int(float(value))
            h, m, s = total // 3600, (total % 3600) // 60, total % 60
    except ValueError:
        return safe_filename(value)
    if h:
        return f"{h:02d}时{m:02d}分{s:02d}秒"
    return f"{m:02d}分{s:02d}秒"


def seconds_to_cn(value: str | None) -> str:
    sec = numeric(value)
    if sec is None:
        return ""
    sec_int = int(sec)
    return f"{sec_int // 60:02d}分{sec_int % 60:02d}秒"


def ensure_unique(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    i = 2
    while True:
        candidate = path.with_name(f"{stem}_{i}{suffix}")
        if not candidate.exists():
            return candidate
        i += 1


def copy_if_exists(src: Path, dst: Path) -> None:
    if src.exists():
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)


def load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("Missing dependency: openpyxl. Use the bundled Codex Python runtime or install openpyxl.") from exc
    return Workbook, LineChart, Reference, Alignment, Border, Font, PatternFill, Side, get_column_letter


def style_sheet(ws, widths: list[int], freeze: str = "A2") -> None:
    _, _, _, Alignment, Border, Font, PatternFill, Side, get_column_letter = load_openpyxl()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    if ws.max_row:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def first_meta(events: list[dict[str, Any]]) -> dict[str, Any]:
    return next((item for item in events if item.get("type") == "session_meta"), {})


def build_context(run_dir: Path, args: argparse.Namespace) -> dict[str, Any]:
    events = read_jsonl(run_dir / "events.jsonl")
    meta = first_meta(events)
    report_dir = run_dir / "report"
    recordings = read_csv(report_dir / "recording_chunks.csv")
    duration = numeric(recordings[0].get("duration_seconds")) if recordings else None
    account = args.account_name or meta.get("account_name") or "未知账号"
    date = args.date or meta.get("session_date") or datetime.now().strftime("%Y-%m-%d")
    if args.duration_label:
        duration_label = args.duration_label
    elif duration:
        duration_label = f"{round(duration / 60):g}分钟"
    else:
        duration_label = "未知时长"
    return {
        "events": events,
        "meta": meta,
        "account": str(account),
        "date": str(date),
        "duration_label": duration_label,
        "recording_duration": duration,
        "report_dir": report_dir,
        "recordings": recordings,
        "metrics": read_csv(report_dir / "metric_samples.csv"),
        "speech": read_csv(report_dir / "speech_segments.csv"),
        "products": read_csv(report_dir / "product_events.csv"),
        "comments": read_csv(report_dir / "comment_events.csv"),
        "orders": read_csv(report_dir / "order_signals.csv"),
        "shots": read_csv(report_dir / "screenshot_events.csv"),
        "loops": read_csv(report_dir / "loop_markers.csv"),
        "transcript": read_csv(run_dir / "raw" / "transcript_segments.csv"),
    }


def default_delivery_dir(run_dir: Path, ctx: dict[str, Any]) -> Path:
    name = f"{ctx['account']}_直播分析_{ctx['date']}_{ctx['duration_label']}"
    return run_dir.parent / safe_filename(name, 120)


def build_screenshot_map(run_dir: Path, delivery_dir: Path, shots: list[dict[str, str]]) -> dict[str, str]:
    shot_dir = delivery_dir / "04_关键截图"
    shot_dir.mkdir(parents=True, exist_ok=True)
    mapping: dict[str, str] = {}
    for shot in shots:
        src_rel = shot.get("file_path") or ""
        src = run_dir / src_rel
        if not src.exists():
            continue
        time_cn = timecode_to_cn(shot.get("recording_timecode"))
        visual = safe_filename(shot.get("visual_state") or shot.get("product_title") or "画面变化", 42)
        dst = ensure_unique(shot_dir / f"{time_cn}_{visual}.png")
        shutil.copy2(src, dst)
        mapping[src_rel] = f"04_关键截图/{dst.name}"
    return mapping


def rel_shot(path: str, screenshot_map: dict[str, str]) -> str:
    if not path:
        return ""
    return screenshot_map.get(path, screenshot_map.get(str(Path(path)), Path(path).name))


def product_sequence(products: list[dict[str, str]]) -> str:
    labels = []
    for item in products:
        title = item.get("product_title") or "未知商品"
        price = item.get("price")
        link = item.get("link_number")
        label = f"{link}号 {title}" if link else title
        if price:
            label += f" {price}元"
        labels.append(label)
    return " -> ".join(labels) if labels else "无可用商品切换记录"


def metric_summary(metrics: list[dict[str, str]]) -> str:
    values = [numeric(item.get("online_viewers")) for item in metrics]
    values = [v for v in values if v is not None]
    if not values:
        return "未采集到在线人数"
    return f"{values[0]:g} 开始，最低 {min(values):g}，最高 {max(values):g}，结束 {values[-1]:g}"


def hot_sale_summary(metrics: list[dict[str, str]]) -> str:
    values = [numeric(item.get("hot_sale_count")) for item in metrics]
    values = [v for v in values if v is not None]
    if len(values) < 2:
        return "热卖数不足以计算变化"
    return f"{values[0]:g} 到 {values[-1]:g}，观察到 {values[-1] - values[0]:+g}"


def build_structured_workbook(delivery_dir: Path, ctx: dict[str, Any], screenshot_map: dict[str, str]) -> None:
    Workbook, LineChart, Reference, *_ = load_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "总览"
    ws.append(["项目", "内容"])
    meta = ctx["meta"]
    rows = [
        ["账号", ctx["account"]],
        ["平台", meta.get("platform", "")],
        ["采集日期", ctx["date"]],
        ["采集时长", f"{ctx.get('recording_duration') or ''} 秒"],
        ["录制范围", meta.get("screen_scope", "")],
        ["采集方法", meta.get("capture_method", "")],
        ["在线人数", metric_summary(ctx["metrics"])],
        ["热卖数", hot_sale_summary(ctx["metrics"])],
        ["商品节奏", product_sequence(ctx["products"])],
        ["循环判断", "未证明完整循环" if not any(item.get("marker_kind") == "probable_completion" for item in ctx["loops"]) else "疑似完整循环"],
        ["备注", "外层为交付文件；技术底稿已收进 _internal_技术底稿。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, [22, 100])

    ws = wb.create_sheet("时间线")
    ws.append(["时间点", "在线人数", "热卖数", "商品/价格", "画面/动作", "对应话术或评论", "证据截图"])
    for metric in ctx["metrics"]:
        shot = next((item for item in ctx["shots"] if item.get("timestamp") == metric.get("timestamp")), {})
        related = [
            sp for sp in ctx["speech"]
            if sp.get("start_time", "") <= metric.get("timestamp", "") <= (sp.get("end_time") or sp.get("start_time", ""))
        ]
        ws.append([
            iso_to_clock(metric.get("timestamp")),
            metric.get("online_viewers"),
            metric.get("hot_sale_count"),
            f"{metric.get('product_title', '')} / {metric.get('product_price', '')}",
            shot.get("visual_state", ""),
            "；".join((sp.get("sales_intent") or sp.get("quote") or "") for sp in related[:2]),
            rel_shot(shot.get("file_path", ""), screenshot_map),
        ])
    style_sheet(ws, [12, 10, 10, 44, 64, 64, 58])

    ws = wb.create_sheet("话术库")
    ws.append(["开始", "结束", "话术类型", "原话/复核稿", "销售意图", "关联商品", "前后在线变化", "后续下单信号数", "证据"])
    for speech in ctx["speech"]:
        delta = numeric(speech.get("online_delta"))
        delta_text = "" if delta is None else f"{speech.get('online_before')} -> {speech.get('online_after')} ({delta:+g})"
        ws.append([
            iso_to_clock(speech.get("start_time")),
            iso_to_clock(speech.get("end_time")),
            SPEECH_TYPE_CN.get(speech.get("speech_type"), speech.get("speech_type", "")),
            speech.get("quote"),
            speech.get("sales_intent"),
            speech.get("product_title"),
            delta_text,
            speech.get("order_signals_after"),
            speech.get("evidence"),
        ])
    style_sheet(ws, [10, 10, 14, 88, 36, 42, 18, 14, 64])

    ws = wb.create_sheet("商品切换")
    ws.append(["时间", "事件", "链接号", "商品", "价格", "话术/卖点", "证据"])
    for product in ctx["products"]:
        ws.append([
            iso_to_clock(product.get("timestamp")),
            EVENT_KIND_CN.get(product.get("event_kind"), product.get("event_kind", "")),
            product.get("link_number"),
            product.get("product_title"),
            product.get("price"),
            product.get("claim_text"),
            rel_shot(product.get("evidence", ""), screenshot_map),
        ])
    style_sheet(ws, [10, 14, 10, 52, 12, 64, 58])

    ws = wb.create_sheet("人数与热卖")
    ws.append(["序号", "时间", "在线人数", "热卖数", "商品", "价格"])
    for idx, metric in enumerate(ctx["metrics"], 1):
        ws.append([
            idx,
            iso_to_clock(metric.get("timestamp")),
            numeric(metric.get("online_viewers")),
            numeric(metric.get("hot_sale_count")),
            metric.get("product_title"),
            numeric(metric.get("product_price")),
        ])
    style_sheet(ws, [8, 12, 12, 12, 52, 12])
    if len(ctx["metrics"]) >= 2:
        chart = LineChart()
        chart.title = "在线人数变化"
        chart.y_axis.title = "在线人数"
        chart.x_axis.title = "时间"
        data = Reference(ws, min_col=3, min_row=1, max_row=len(ctx["metrics"]) + 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=len(ctx["metrics"]) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 18
        ws.add_chart(chart, "H2")

    ws = wb.create_sheet("评论问题")
    ws.append(["时间", "评论类型", "评论/问题", "关联商品", "是否被回应", "证据"])
    for comment in ctx["comments"]:
        ws.append([
            iso_to_clock(comment.get("timestamp")),
            COMMENT_TYPE_CN.get(comment.get("comment_type"), comment.get("comment_type", "")),
            comment.get("comment_text"),
            comment.get("related_product_id"),
            comment.get("response_observed"),
            comment.get("evidence"),
        ])
    style_sheet(ws, [10, 16, 82, 14, 12, 64])

    ws = wb.create_sheet("下单信号")
    ws.append(["时间", "信号类型", "商品", "变化前", "变化后", "说明", "是否平台验证订单", "证据"])
    for order in ctx["orders"]:
        ws.append([
            iso_to_clock(order.get("timestamp")),
            order.get("signal_kind"),
            order.get("product_title"),
            order.get("value_before"),
            order.get("value_after"),
            order.get("signal_text"),
            order.get("verified_order_count"),
            rel_shot(order.get("evidence", ""), screenshot_map),
        ])
    style_sheet(ws, [10, 18, 52, 10, 10, 56, 16, 58])

    ws = wb.create_sheet("截图索引")
    ws.append(["时间码", "截图文件", "截图原因", "画面内容", "商品"])
    for shot in ctx["shots"]:
        ws.append([
            shot.get("recording_timecode"),
            rel_shot(shot.get("file_path", ""), screenshot_map),
            shot.get("screenshot_reason"),
            shot.get("visual_state"),
            shot.get("product_title"),
        ])
    style_sheet(ws, [12, 62, 18, 84, 50])

    ws = wb.create_sheet("循环判断")
    ws.append(["判断项", "内容"])
    loop_complete = any(item.get("marker_kind") == "probable_completion" for item in ctx["loops"])
    ws.append(["结论", "已标记完整循环" if loop_complete else "未证明完整循环"])
    ws.append(["证据数量", f"{len(ctx['loops'])} 条 loop_marker"])
    ws.append(["建议", "长时间竞品监控建议每 30-60 分钟分段，直到完整循环被验证。"])
    style_sheet(ws, [20, 100])

    wb.save(delivery_dir / "03_结构化数据总表.xlsx")


def build_transcript_workbook(delivery_dir: Path, ctx: dict[str, Any]) -> None:
    Workbook, *_ = load_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "逐字稿"
    ws.append(["序号", "开始时间", "结束时间", "自动逐字稿", "识别可信度提示", "备注"])
    for idx, row in enumerate(ctx["transcript"], 1):
        no_speech = numeric(row.get("no_speech_prob")) or 0
        ws.append([
            idx,
            seconds_to_cn(row.get("start")),
            seconds_to_cn(row.get("end")),
            row.get("text"),
            "需要复核" if no_speech > 0.6 else "可参考",
            "自动转写，未逐句人工精校",
        ])
    style_sheet(ws, [8, 12, 12, 110, 16, 28])

    ws = wb.create_sheet("高价值话术提炼")
    ws.append(["开始", "结束", "类型", "复核后的关键话术", "销售意图", "关联商品"])
    for speech in ctx["speech"]:
        ws.append([
            iso_to_clock(speech.get("start_time")),
            iso_to_clock(speech.get("end_time")),
            SPEECH_TYPE_CN.get(speech.get("speech_type"), speech.get("speech_type", "")),
            speech.get("quote"),
            speech.get("sales_intent"),
            speech.get("product_title"),
        ])
    style_sheet(ws, [10, 10, 14, 92, 38, 46])
    wb.save(delivery_dir / "02_直播逐字稿_带时间轴.xlsx")


def write_markdown(delivery_dir: Path, ctx: dict[str, Any]) -> None:
    speech_types = Counter(SPEECH_TYPE_CN.get(item.get("speech_type"), item.get("speech_type", "未知")) for item in ctx["speech"])
    top_speech = "、".join(f"{key} {count}条" for key, count in speech_types.most_common()) or "暂无"
    first = f"""# 本次结论｜{ctx['account']}直播分析（{ctx['duration_label']}）

## 先看结论

本次样本已整理成交付版目录。外层文件只保留完整录屏、逐字稿、结构化总表、关键截图和分析报告；技术文件已收进 `_internal_技术底稿`。

## 本次数据口径

- 采集对象：{ctx['meta'].get('platform', '')}，{ctx['account']}
- 采集日期：{ctx['date']}
- 录制时长：{ctx.get('recording_duration') or ''} 秒
- 录制范围：{ctx['meta'].get('screen_scope', '')}
- 采集方法：{ctx['meta'].get('capture_method', '')}

## 关键数字

- 在线人数：{metric_summary(ctx['metrics'])}
- 热卖数：{hot_sale_summary(ctx['metrics'])}
- 商品节奏：{product_sequence(ctx['products'])}
- 话术类型：{top_speech}

## 最值得看的文件

1. `01_完整录屏_{safe_filename(ctx['account'])}_{ctx['date']}_{ctx['duration_label']}.mov`
2. `02_直播逐字稿_带时间轴.xlsx`
3. `03_结构化数据总表.xlsx`
4. `05_话术分析报告.md`
5. `04_关键截图/`
"""
    delivery_dir.joinpath("00_先看这个_本次结论.md").write_text(first, encoding="utf-8")

    quote_lines = []
    for item in ctx["speech"][:12]:
        quote_lines.append(f"- **{SPEECH_TYPE_CN.get(item.get('speech_type'), item.get('speech_type', ''))}**：{item.get('quote', '')}")
    relation_lines = []
    for item in ctx["speech"][:8]:
        delta = numeric(item.get("online_delta"))
        delta_text = "" if delta is None else f"，在线人数 {item.get('online_before')} -> {item.get('online_after')} ({delta:+g})"
        relation_lines.append(f"- {iso_to_clock(item.get('start_time'))} {SPEECH_TYPE_CN.get(item.get('speech_type'), item.get('speech_type', ''))}：{item.get('sales_intent', '')}{delta_text}，后续下单信号 {item.get('order_signals_after', '')} 个。")
    analysis = f"""# 话术分析报告｜{ctx['account']}

## 核心打法

本报告基于录屏、逐字稿、截图、在线人数、评论问题和可见下单信号整理。下方判断只说明样本内的时间关系；除非多轮重复验证，不把相关性写成因果。

## 主要话术模块

{chr(10).join(quote_lines) if quote_lines else '- 暂无结构化话术。'}

## 话术与数据关系

{chr(10).join(relation_lines) if relation_lines else '- 暂无可计算关系。'}

## 商品节奏

{product_sequence(ctx['products'])}

## 循环判断

{'已出现完整循环标记。' if any(item.get('marker_kind') == 'probable_completion' for item in ctx['loops']) else '本次样本没有证明完整循环。若用于长时间竞品监控，应继续分段录制直到循环闭合。'}

## 可借鉴

- 把时间、发货、赠品、库存和链接选择讲成可执行的购买理由。
- 把评论区高频疑问沉淀为固定答法，尤其是发货、品质、规格差异、礼品场景。
- 用商品切换承接不同人群需求，而不是只反复推一个链接。

## 不建议照搬

- 自动转写未经逐句精校，引用对外材料前需要人工复核。
- 单个 10 分钟样本不足以判断强因果，需要更长监控和多轮同类片段对比。
"""
    delivery_dir.joinpath("05_话术分析报告.md").write_text(analysis, encoding="utf-8")


def copy_user_facing_files(run_dir: Path, delivery_dir: Path, ctx: dict[str, Any], screenshot_map: dict[str, str]) -> None:
    raw = run_dir / "raw"
    delivery_dir.mkdir(parents=True, exist_ok=True)
    recording = ctx["recordings"][0] if ctx["recordings"] else {}
    recording_path = run_dir / (recording.get("file_path") or "raw/screen-recording-001.mov")
    if recording_path.exists():
        dst_name = f"01_完整录屏_{safe_filename(ctx['account'])}_{ctx['date']}_{ctx['duration_label']}{recording_path.suffix}"
        shutil.copy2(recording_path, delivery_dir / dst_name)
    copy_if_exists(raw / "screenshots-contact-sheet.jpg", delivery_dir / "06_截图索引图.jpg")


def copy_internal_files(run_dir: Path, delivery_dir: Path) -> None:
    internal = delivery_dir / "_internal_技术底稿"
    internal.mkdir(parents=True, exist_ok=True)
    copy_if_exists(run_dir / "events.jsonl", internal / "events.jsonl")
    copy_if_exists(run_dir / "assumptions.md", internal / "采集说明.md")
    copy_if_exists(run_dir / "raw" / "capture-target-check.png", internal / "窗口录制范围验证.png")
    copy_if_exists(run_dir / "raw" / "audio-source-check.txt", internal / "音频有效性检查.txt")
    copy_if_exists(run_dir / "raw" / "transcript_raw.json", internal / "逐字稿原始json.json")
    if (run_dir / "report").exists():
        target = internal / "原始报表csv"
        if target.exists():
            shutil.rmtree(target)
        shutil.copytree(run_dir / "report", target)


def main() -> None:
    args = parse_args()
    run_dir = Path(args.run_dir).expanduser().resolve()
    if not run_dir.exists():
        raise SystemExit(f"Run directory not found: {run_dir}")
    ctx = build_context(run_dir, args)
    delivery_dir = Path(args.out_dir).expanduser().resolve() if args.out_dir else default_delivery_dir(run_dir, ctx).resolve()
    delivery_dir.mkdir(parents=True, exist_ok=True)

    screenshot_map = build_screenshot_map(run_dir, delivery_dir, ctx["shots"])
    copy_user_facing_files(run_dir, delivery_dir, ctx, screenshot_map)
    build_transcript_workbook(delivery_dir, ctx)
    build_structured_workbook(delivery_dir, ctx, screenshot_map)
    write_markdown(delivery_dir, ctx)
    copy_internal_files(run_dir, delivery_dir)
    print(f"Wrote delivery package to {delivery_dir}")


if __name__ == "__main__":
    main()
